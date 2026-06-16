#!/usr/bin/env python3
import argparse
import datetime as dt
import json
import math
import os
import re
import sqlite3
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit("Missing dependency: openpyxl is required to read .xlsx color-card files.") from exc


ROOT_DIR = Path(__file__).resolve().parents[2]
SKILL_DIR = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = ROOT_DIR / "MGS-PET改色膜-飞书色卡画廊_色卡档案_全部色卡.xlsx"
DEFAULT_PREVIEW_DIR = SKILL_DIR / "assets" / "previews"
DEFAULT_OUT = SKILL_DIR / "references" / "color_assets.json"
DEFAULT_CARD_DB = Path("/Users/fkycoya/Documents/Card-color/data/color-lab.db")


def normalize_text(value) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split()).strip()


def slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def prompt_finish(raw_finish: str, material: str) -> tuple[str, str]:
    finish = raw_finish or ""
    material_label = material or "车膜"
    if any(token in finish for token in ("哑", "matte", "Matte")):
        return finish, f"{material_label}哑光膜面，低反射、柔和高光"
    if any(token in finish for token in ("亮", "gloss", "Gloss")):
        return finish, f"{material_label}亮光膜面，清晰环境反射与高光"
    if any(token in finish for token in ("金属", "metal", "Metal")):
        return finish, f"{material_label}金属膜面，细腻金属颗粒与车身曲面高光"
    if any(token in finish for token in ("珠光", "pearl", "Pearl")):
        return finish, f"{material_label}珠光膜面，柔和珠光层次"
    return finish or "未标注", f"{material_label}改色膜，膜面效果以 preview 色卡图为准，HEX/Lab 仅作辅助标注"


def find_preview(serial: str, source_name: str, preview_dir: Path) -> tuple[str, str]:
    if not preview_dir.is_dir():
        return "", "missing_preview_dir"

    exact = preview_dir / source_name if source_name else None
    if exact and exact.is_file():
        return exact.as_posix(), "exact_source_name"

    matches = sorted(preview_dir.glob(f"{serial}-measurement-*.png"))
    if matches:
        return matches[0].as_posix(), "serial_prefix"

    return "", "not_found"


def float_or_none(value):
    if value is None:
        return None
    try:
        if math.isnan(float(value)):
            return None
    except TypeError:
        return None
    return round(float(value), 2)


def normalize_hex(value: str) -> str:
    text = normalize_text(value).upper()
    if not text:
        return ""
    if not text.startswith("#"):
        text = f"#{text}"
    if not re.fullmatch(r"#[0-9A-F]{6}", text):
        raise ValueError(f"Invalid HEX color value: {value}")
    return text


def load_card_hex_values(card_db: Path) -> dict[str, dict]:
    if not card_db.is_file():
        return {}

    conn = sqlite3.connect(card_db)
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            """
            SELECT
              color_code,
              lab_l,
              lab_a,
              lab_b,
              preview_hex,
              preview_r,
              preview_g,
              preview_b,
              preview_png_path,
              measured_at,
              id
            FROM lab_measurements
            ORDER BY color_code, measured_at DESC, id DESC
            """
        ).fetchall()
    finally:
        conn.close()

    latest: dict[str, dict] = {}
    for row in rows:
        code = normalize_text(row["color_code"]).upper()
        if not code or code in latest:
            continue
        latest[code] = {
            "hex": normalize_hex(row["preview_hex"]),
            "lab": {
                "L": float_or_none(row["lab_l"]),
                "a": float_or_none(row["lab_a"]),
                "b": float_or_none(row["lab_b"]),
            },
            "preview_png_path": normalize_text(row["preview_png_path"]),
            "measured_at": normalize_text(row["measured_at"]),
            "measurement_id": row["id"],
        }
    return latest


def build_assets(source: Path, preview_dir: Path, out_path: Path, card_db: Path | None = None) -> dict:
    wb = openpyxl.load_workbook(source, data_only=True)
    ws = wb.active
    card_hex_values = load_card_hex_values(card_db) if card_db else {}
    headers = [normalize_text(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = dict(zip(headers, row))
        serial = normalize_text(item.get("序列号"))
        zh_name = normalize_text(item.get("中文名称"))
        en_name = normalize_text(item.get("英文名称"))
        if not serial or not zh_name:
            continue

        l_value = float_or_none(item.get("L 值"))
        a_value = float_or_none(item.get("a 值"))
        b_value = float_or_none(item.get("b 值"))
        if l_value is None or a_value is None or b_value is None:
            continue

        material = normalize_text(item.get("材料")) or "PET"
        raw_finish = normalize_text(item.get("表面效果"))
        finish, finish_hint = prompt_finish(raw_finish, material)
        family = normalize_text(item.get("色系"))
        source_swatch = normalize_text(item.get("色块图片"))
        card_hex = card_hex_values.get(serial.upper())
        if card_hex and card_hex.get("preview_png_path"):
            source_swatch = Path(card_hex["preview_png_path"]).name
        preview_path, match_strategy = find_preview(serial, source_swatch, preview_dir)
        preview_rel = ""
        if preview_path:
            preview_rel = os.path.relpath(Path(preview_path).resolve(), out_path.parent.resolve())

        lab_value = f"Lab(L={l_value}, a={a_value}, b={b_value})"
        hex_value = card_hex["hex"] if card_hex else ""
        color_value = hex_value or lab_value
        description_parts = []
        if family:
            description_parts.append(family)
        description_parts.append(finish_hint)
        description_parts.append(f"{zh_name}（{en_name}）")
        description_parts.append("以随图传入的 preview 色卡图作为目标颜色主依据")
        if hex_value:
            description_parts.append(f"HEX 辅助标注 {hex_value}")
        description_parts.append(f"Lab 辅助参考 {lab_value}")

        rows.append(
            {
                "id": serial,
                "source_id": normalize_text(item.get("ID")),
                "serial": serial,
                "names": {
                    "zh": zh_name,
                    "en": en_name,
                    "aliases": [value for value in [serial, zh_name, en_name, slug(zh_name), slug(en_name)] if value],
                },
                "color": {
                    "lab": {"L": l_value, "a": a_value, "b": b_value},
                    "hex": hex_value or None,
                    "family": family or None,
                },
                "material": material,
                "finish": {
                    "raw": raw_finish or None,
                    "prompt_label": finish,
                    "prompt_hint": finish_hint,
                },
                "images": {
                    "swatch": preview_rel or None,
                    "source_swatch_name": source_swatch or None,
                    "real_card": normalize_text(item.get("真实色卡图片")) or None,
                    "match_strategy": match_strategy,
                },
                "measurement": {
                    "source": "Card-color SQLite preview_hex" if card_hex else "Lab workbook only",
                    "card_db": card_db.as_posix() if card_db and card_db.is_file() else None,
                    "measurement_id": card_hex.get("measurement_id") if card_hex else None,
                    "measured_at": card_hex.get("measured_at") if card_hex else None,
                    "preview_lab": card_hex.get("lab") if card_hex else None,
                },
                "prompt": {
                    "color_name": f"{zh_name} / {en_name}",
                    "color_code": serial,
                    "color_value": color_value,
                    "lab_value": lab_value,
                    "reference_priority": "preview_swatch_image",
                    "finish": finish_hint,
                    "description": "，".join(description_parts),
                },
                "tags": [value for value in [material, family, raw_finish, zh_name, en_name] if value],
            }
        )

    return {
        "schema_version": "1.4.0",
        "asset_type": "vehicle_wrap_color_library",
        "generated_at": dt.datetime.now(dt.UTC).isoformat(timespec="seconds"),
            "source": {
                "workbook": source.name,
                "sheet": ws.title,
                "preview_dir": preview_dir.as_posix(),
                "card_db": card_db.as_posix() if card_db and card_db.is_file() else None,
                "lab_observer": "CIE Lab from measured color-card data",
                "color_value_policy": "Use the preview swatch image as the primary generation color reference. HEX and Lab are auxiliary labels for routing, traceability, and sales summaries.",
            },
        "usage": {
            "lookup_keys": ["id", "serial", "names.zh", "names.en", "names.aliases"],
            "generation_script": "scripts/gen.py --asset-id <serial-or-name> --vehicle-ref <customer-car-image>",
        },
        "counts": {
            "assets": len(rows),
            "with_swatch": sum(1 for item in rows if item["images"]["swatch"]),
            "with_hex": sum(1 for item in rows if item["color"].get("hex")),
        },
        "assets": rows,
    }


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(description="Build the car wrap color asset JSON library from the Lab color-card workbook.")
    ap.add_argument("--source", default=DEFAULT_SOURCE.as_posix(), help="Source .xlsx color-card workbook.")
    ap.add_argument("--preview-dir", default=DEFAULT_PREVIEW_DIR.as_posix(), help="Directory containing generated swatch images stored inside this skill.")
    ap.add_argument("--out", default=DEFAULT_OUT.as_posix(), help="Output JSON asset library path.")
    ap.add_argument(
        "--card-db",
        default=os.environ.get("CARD_COLOR_DB", DEFAULT_CARD_DB.as_posix()),
        help="Card-color SQLite database containing authoritative preview_hex values.",
    )
    return ap.parse_args()


def main() -> int:
    args = parse_args()
    source = Path(args.source).expanduser().resolve()
    preview_dir = Path(args.preview_dir).expanduser().resolve()
    out_path = Path(args.out).expanduser().resolve()
    card_db = Path(args.card_db).expanduser().resolve() if args.card_db else None
    if not source.is_file():
        raise FileNotFoundError(f"Source workbook not found: {source}")
    library = build_assets(source, preview_dir, out_path, card_db)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(library, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({"out": out_path.as_posix(), "counts": library["counts"]}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
